#!/usr/bin/env python

""" Converts a JPEG image into an excel spreadsheet, where each cell is a pixel. """

import os, sys

from PIL import Image

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill

def rgb_to_hex(r: int, g: int, b:int) -> str:
    """ Converts a set of RGB integer values into a HEX-RGB string equivalent.
    
    Parameters
    ----------
    r : int
        Red color value. Must be an int in range(0, 256).
    g : int
        Green color value. Must be an int in range(0, 256).
    b : int
        Blue color value. Must be an int in range(0, 256).

    Returns
    -------
    output : str
        String representation of the RGB color value in the format '#ABCDEF'

    """
    for input_value in (r, g, b):
        # make sure input values are valid
        if type(input_value) != int:
            raise TypeError("Expected type int.")
        if input_value not in range(0,256):
            raise ValueError("Expected value in range(0, 255) but got {}.".format(str(input_value)))

    # convert to hex
    output = '#{:02x}{:02x}{:02x}'.format(r, g, b)

    return output



def main() -> None:
    """ Converts a JPEG image into an excel spreadsheet, where each cell is a pixel.
        Attempts to use the file specified by the first commandline argument as the
        input picture.

        XXX Not tested for greyscale/rgba images. will probably crash for those
    """

    # image gets downscaled to this resolution as to not generate too many
    # spreadsheet cells
    max_image_width = 200
    max_image_height = 200

    # size of resulting spreadsheet cells. Works for LibreOffice Calc
    xlsx_pixel_height = 7
    xlsx_pixel_width = xlsx_pixel_height / 7

    # make sure exactly one argument was supplied
    if len(sys.argv) != 2:
        print("ERROR: Missing argument.")
        sys.exit(1)

    path_to_input_image = sys.argv[1]

    # make sure argument is a file
    if not os.path.isfile(path_to_input_image):
        print("ERROR: {} is not a valid file.".format(path_to_input_image))
        sys.exit(1)

    # make sure image is readable
    if not os.access(path_to_input_image, os.R_OK):
        print("ERROR: {} is not readable.".format(path_to_input_image))
        sys.exit(1)

    # load image
    try:
        image = Image.open(path_to_input_image)
    except OSError:
        print("ERROR: {} is not an image or cannot be opened.".format(path_to_input_image))
        sys.exit(1)

    # determine image title and discard file extension
    image_title = os.path.basename(path_to_input_image)
    if '.' in image_title:
        image_title = image_title.rsplit(sep='.', maxsplit=1)[0]

    # downscale image if necessary
    image.thumbnail((max_image_width, max_image_height), Image.ANTIALIAS)
    image_width  = image.size[0]
    image_height = image.size[1]

    # create an excel worksheet
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = image_title
    xlsx_file = image_title + '.xlsx'

    # insert image's pixels into the spreadsheet
    for x in range(image_width):
        for y in range(image_height):
            # spreadsheet indexing starts at 1
            column = x + 1
            row    = y + 1

            # retrieve pixel's color value
            pixel_color_rgb = image.getpixel((x, y))
            pixel_color_hex = rgb_to_hex(pixel_color_rgb[0], pixel_color_rgb[1], pixel_color_rgb[2])

            # color in the corresponding cell
            cell = worksheet.cell(column=column, row=row)
            cell.fill = PatternFill(start_color=pixel_color_hex[1:], fill_type = "solid")

            # resize cell
            worksheet.column_dimensions[get_column_letter(column)].width = xlsx_pixel_width
            worksheet.row_dimensions[row].height                         = xlsx_pixel_height

    # Save the spreadsheet
    workbook.save(xlsx_file)

if __name__ == '__main__':
    main()
